# -*- coding: utf-8 -*- 
# @Time : 2022-05-29 15:11 
# @Author : kunp
# @File : ExcelUtil.py 
# @Software: PyCharm
from openpyxl import load_workbook

# from Conf.VarConfig import testDataPath


class ExcelUtil:

    def __init__(self):
        self.workbook =None
        # self.sheet = None

    def load_workbook(self,filename):
        """
        加载名为 filename 的文件
        :param filename:
        :return:
        """
        try:
            self.workbook = load_workbook(filename)
        except Exception as e:
            print(e)

    def get_sheet_by_name(self,sheetname):
        """
        通过 sheetname 读取 sheet （页）
        :param sheetname:
        :return:
        """
        try:
            return  self.workbook[sheetname]
        except Exception as e:
            print(e)

    def get_row_nums(self, sheet):
        """
        返回当前页的行数
        :return:
        """
        return sheet.max_row

    def get_col_nums(self, sheet):
        """
        返回当前页的列数
        :return:
        """
        return sheet.max_column


    def get_row_values(self,row, sheet):
        """
        获取某一行的所有值，返回个列表
        :param row:
        :return:
        """
        columns = sheet.max_column
        row_data = []
        for i in range(1,columns+1):
            cell_value = sheet.cell(row=row,column=i).value
            row_data.append(cell_value)
        return row_data

    def get_all_rows_values(self, max_row, sheet):
        """
        获取所有行的所以值，返回个列表
        :param max_row:
        :param cloumn:
        :param value:
        :param filename:
        :return:
        """
        columns = sheet.max_column

        row_1 = []
        for i1 in range(1, columns + 1):
            row_1.append(sheet.cell(row=1,column=i1).value)
        # 用列表记录每行数据
        dic_data = []
        for j in range(2,max_row+1):
            # 获取各行数据，并且用列表记录，每次循环充值空序列s
            row_data = []
            for i in range(1,columns+1):
                row_data.append(sheet.cell(row=j,column=i).value)

            #将数据zip序列化
            # row_1_copy = row_1.copy()
            # row_data_copy = row_data.copy()
            data = zip(row_1,row_data)
            #进行字典化后，存入列表
            coco = dict(data).copy()

            dic_data.append(coco)
        return dic_data


    def get_cell_value(self,row,column,sheet):
        """
        获取某一单元格的值
        :param row:
        :param column:
        :return:
        """
        return sheet.cell(row=row,column=column).value


    def write_cell(self,row, column ,value, filename, sheet):
        """
        写入某个值
        :param row:
        :param cloumn:
        :param value:
        :param filename:
        :return:
        """
        sheet.cell(row=row, column=column, value=value)
        self.workbook.save(filename)


if __name__ == '__main__':
    filename = testDataPath + "/test.xlsx"
    excel = ExcelUtil()
    excel.load_workbook(filename)
    excel.get_sheet_by_name("login")
    # print(excel.get_row_nums())
    # print(excel.get_col_nums())
    headers = excel.get_row_values(1)
    values = excel.get_row_values(2)

    print(dict(zip(headers,values)))
    # print(excel.get_cell_value(1,1))

    # lst = [1, '登录正常流程', 'cc', 123, 'Y', None]
    # username = lst[2]
    # password = lst[3]
    #
    # dic = {'id':1,'usenrmae':"cc",'pws':'123'...}
    # username = dic['usenrmae']
    # pwd = dic['pws']
